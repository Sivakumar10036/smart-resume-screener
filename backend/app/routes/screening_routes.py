from datetime import datetime, timezone
from io import BytesIO

from bson import ObjectId

from fastapi import (
    APIRouter,
    Depends,
    HTTPException
)

from fastapi.responses import StreamingResponse

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side
)

from app.config import settings

from app.database import (
    candidates_collection,
    jobs_collection,
    screening_results_collection
)

from app.schemas.screening_schema import (
    BatchScreeningRequest
)

from app.services.matcher import (
    match_resume_with_job
)

from app.utils.security import (
    get_current_user,
    require_roles
)


router = APIRouter(
    prefix="/api/screening",
    tags=["Screening"]
)


def normalize_score(
    value
):

    try:

        score = float(
            value
        )

    except (
        TypeError,
        ValueError
    ):

        score = 0


    if score < 0:

        score = 0


    if score <= 10:

        score = score * 10


    if score > 100:

        score = 100


    return round(
        score,
        1
    )


def get_recommendation(score):

    score = float(score)

    if score >= settings.shortlist_threshold:
        return "SHORTLIST"

    elif score >= settings.review_threshold:
        return "REVIEW"

    else:
        return "REJECT"


def get_user_id(
    current_user
):

    return current_user["_id"]


def get_candidate_owner_id(
    candidate
):

    owner_id = candidate.get(
        "owner_id"
    )


    if owner_id:

        return owner_id


    uploaded_by = candidate.get(
        "uploaded_by"
    )


    if uploaded_by:

        return uploaded_by


    return None


def candidate_belongs_to_user(
    candidate,
    current_user
):

    owner_id = get_candidate_owner_id(
        candidate
    )


    if owner_id is None:

        return False


    current_user_id = str(
        get_user_id(
            current_user
        )
    )


    return (
        str(owner_id)
        ==
        current_user_id
    )


def format_candidate_result(
    result,
    candidate,
    job,
    rank=None
):

    score = normalize_score(
        result.get(
            "match_score",
            0
        )
    )


    formatted_result = {

        "result_id":
            str(
                result["_id"]
            ),

        "candidate_id":
            str(
                result["candidate_id"]
            ),

        "candidate_name":
            candidate.get(
                "name",
                "Unknown Candidate"
            )
            if candidate
            else "Unknown Candidate",

        "candidate_email":
            candidate.get(
                "email",
                ""
            )
            if candidate
            else "",

        "candidate_phone":
            candidate.get(
                "phone",
                ""
            )
            if candidate
            else "",

        "job_id":
            str(
                result["job_id"]
            ),

        "job_title":
            job.get(
                "title",
                ""
            )
            if job
            else "",

        "match_score":
            score,

        "recommendation":
            get_recommendation(
                score
            ),

        "matched_skills":
            result.get(
                "matched_skills",
                []
            ),

        "missing_skills":
            result.get(
                "missing_skills",
                []
            ),

        "strengths":
            result.get(
                "strengths",
                []
            ),

        "justification":
            result.get(
                "justification",
                ""
            ),

        "created_at":
            result.get(
                "created_at"
            )
    }


    if rank is not None:

        formatted_result["rank"] = rank


    return formatted_result


def create_screening_document(
    candidate,
    job,
    result,
    current_user
):

    score = normalize_score(
        result.get(
            "match_score",
            0
        )
    )


    result["match_score"] = score


    recommendation = get_recommendation(
        score
    )


    owner_id = get_candidate_owner_id(
        candidate
    )


    return {

        "candidate_id":
            candidate["_id"],

        "candidate_owner_id":
            str(owner_id)
            if owner_id is not None
            else None,

        "job_id":
            job["_id"],

        "match_score":
            score,

        "matched_skills":
            result.get(
                "matched_skills",
                []
            ),

        "missing_skills":
            result.get(
                "missing_skills",
                []
            ),

        "strengths":
            result.get(
                "strengths",
                []
            ),

        "recommendation":
            recommendation,

        "justification":
            result.get(
                "justification",
                ""
            ),

        "screened_by":
            str(
                current_user["_id"]
            ),

        "screened_by_username":
            current_user.get(
                "username",
                ""
            ),

        "created_at":
            datetime.now(
                timezone.utc
            )
    }


@router.post("/my-match")
def match_my_resume(

    candidate_id: str,

    job_id: str,

    current_user=Depends(
        require_roles(
            "VIEWER"
        )
    )
):

    try:

        candidate = (
            candidates_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            candidate_id
                        )
                }
            )
        )


        job = (
            jobs_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            job_id
                        ),

                    "is_active":
                        {
                            "$ne":
                                False
                        }
                }
            )
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid resume or job ID"
        )


    if not candidate:

        raise HTTPException(
            status_code=404,
            detail="Resume not found"
        )


    if not job:

        raise HTTPException(
            status_code=404,
            detail="Job not found"
        )


    if not candidate_belongs_to_user(
        candidate,
        current_user
    ):

        raise HTTPException(
            status_code=403,
            detail="You can only screen your own resumes"
        )


    result = match_resume_with_job(
        candidate,
        job
    )


    screening_document = (
        create_screening_document(
            candidate,
            job,
            result,
            current_user
        )
    )


    database_result = (
        screening_results_collection
        .insert_one(
            screening_document
        )
    )


    score = screening_document[
        "match_score"
    ]


    recommendation = screening_document[
        "recommendation"
    ]


    result["match_score"] = score

    result["recommendation"] = (
        recommendation
    )


    return {

        "message":
            "Your resume screening completed",

        "screening_id":
            str(
                database_result.inserted_id
            ),

        "candidate_id":
            str(
                candidate["_id"]
            ),

        "candidate":
            candidate.get(
                "name",
                ""
            ),

        "job":
            job.get(
                "title",
                ""
            ),

        "job_title":
            job.get(
                "title",
                ""
            ),

        "result":
            result
    }


@router.post("/match")
def match_candidate(

    candidate_id: str,

    job_id: str,

    current_user=Depends(
        get_current_user
    )
):

    try:

        candidate = (
            candidates_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            candidate_id
                        )
                }
            )
        )


        job = (
            jobs_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            job_id
                        ),

                    "is_active":
                        {
                            "$ne":
                                False
                        }
                }
            )
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid candidate or job ID"
        )


    if not candidate:

        raise HTTPException(
            status_code=404,
            detail="Candidate not found"
        )


    if not job:

        raise HTTPException(
            status_code=404,
            detail="Job description not found"
        )


    user_role = str(
        current_user.get(
            "role",
            "VIEWER"
        )
    ).upper()


    if user_role == "VIEWER":

        if not candidate_belongs_to_user(
            candidate,
            current_user
        ):

            raise HTTPException(
                status_code=403,
                detail="You can only screen your own resumes"
            )


    elif user_role not in [
        "ADMIN",
        "RECRUITER"
    ]:

        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )


    result = match_resume_with_job(
        candidate,
        job
    )


    screening_document = (
        create_screening_document(
            candidate,
            job,
            result,
            current_user
        )
    )


    database_result = (
        screening_results_collection
        .insert_one(
            screening_document
        )
    )


    score = screening_document[
        "match_score"
    ]


    recommendation = screening_document[
        "recommendation"
    ]


    result["match_score"] = score

    result["recommendation"] = (
        recommendation
    )


    return {

        "message":
            "Candidate screening completed",

        "screening_id":
            str(
                database_result.inserted_id
            ),

        "candidate_id":
            str(
                candidate["_id"]
            ),

        "candidate":
            candidate.get(
                "name",
                ""
            ),

        "job":
            job.get(
                "title",
                ""
            ),

        "job_title":
            job.get(
                "title",
                ""
            ),

        "result":
            result
    }


@router.post("/batch")
def batch_screen_candidates(

    request: BatchScreeningRequest,

    current_user=Depends(
        require_roles(
            "ADMIN",
            "RECRUITER"
        )
    )
):

    try:

        job = (
            jobs_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            request.job_id
                        )
                }
            )
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid job ID"
        )


    if not job:

        raise HTTPException(
            status_code=404,
            detail="Job description not found"
        )


    ranked_candidates = []


    for candidate_id in request.candidate_ids:

        try:

            candidate = (
                candidates_collection
                .find_one(
                    {
                        "_id":
                            ObjectId(
                                candidate_id
                            )
                    }
                )
            )

        except Exception:

            continue


        if not candidate:

            continue


        result = match_resume_with_job(
            candidate,
            job
        )


        screening_document = (
            create_screening_document(
                candidate,
                job,
                result,
                current_user
            )
        )


        database_result = (
            screening_results_collection
            .insert_one(
                screening_document
            )
        )


        score = screening_document[
            "match_score"
        ]


        recommendation = screening_document[
            "recommendation"
        ]


        ranked_candidates.append({

            "result_id":
                str(
                    database_result.inserted_id
                ),

            "candidate_id":
                str(
                    candidate["_id"]
                ),

            "candidate_name":
                candidate.get(
                    "name",
                    "Unknown"
                ),

            "candidate_email":
                candidate.get(
                    "email",
                    ""
                ),

            "candidate_phone":
                candidate.get(
                    "phone",
                    ""
                ),

            "candidate_skills":
                candidate.get(
                    "skills",
                    []
                ),

            "candidate_education":
                candidate.get(
                    "education",
                    []
                ),

            "candidate_experience":
                candidate.get(
                    "experience",
                    []
                ),

            "match_score":
                score,

            "recommendation":
                recommendation,

            "matched_skills":
                result.get(
                    "matched_skills",
                    []
                ),

            "missing_skills":
                result.get(
                    "missing_skills",
                    []
                ),

            "strengths":
                result.get(
                    "strengths",
                    []
                ),

            "justification":
                result.get(
                    "justification",
                    ""
                )
        })


    ranked_candidates.sort(
        key=lambda candidate:
            candidate["match_score"],
        reverse=True
    )


    for index, candidate in enumerate(
        ranked_candidates,
        start=1
    ):

        candidate["rank"] = index


    shortlisted_count = sum(

        1

        for candidate
        in ranked_candidates

        if candidate[
            "recommendation"
        ] == "SHORTLIST"
    )


    return {

        "message":
            "Batch screening completed",

        "job_id":
            request.job_id,

        "job_title":
            job.get(
                "title",
                ""
            ),

        "total_candidates":
            len(
                ranked_candidates
            ),

        "shortlisted":
            shortlisted_count,

        "candidates":
            ranked_candidates
    }


@router.post("/calculate/{job_id}")
def calculate_screening_results(

    job_id: str,

    current_user=Depends(
        require_roles(
            "ADMIN",
            "RECRUITER"
        )
    )
):

    try:

        job_object_id = ObjectId(
            job_id
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid job ID"
        )


    job = (
        jobs_collection
        .find_one(
            {
                "_id":
                    job_object_id
            }
        )
    )


    if not job:

        raise HTTPException(
            status_code=404,
            detail="Job description not found"
        )


    candidates = list(
        candidates_collection.find({})
    )


    if not candidates:

        raise HTTPException(
            status_code=404,
            detail="No candidates found in the database"
        )


    screening_results_collection.delete_many(
        {
            "job_id":
                job_object_id
        }
    )


    ranked_candidates = []


    for candidate in candidates:

        result = match_resume_with_job(
            candidate,
            job
        )


        screening_document = (
            create_screening_document(
                candidate,
                job,
                result,
                current_user
            )
        )


        database_result = (
            screening_results_collection
            .insert_one(
                screening_document
            )
        )


        score = screening_document[
            "match_score"
        ]


        recommendation = screening_document[
            "recommendation"
        ]


        ranked_candidates.append({

            "result_id":
                str(
                    database_result.inserted_id
                ),

            "candidate_id":
                str(
                    candidate["_id"]
                ),

            "candidate_name":
                candidate.get(
                    "name",
                    "Unknown"
                ),

            "candidate_email":
                candidate.get(
                    "email",
                    ""
                ),

            "candidate_phone":
                candidate.get(
                    "phone",
                    ""
                ),

            "candidate_skills":
                candidate.get(
                    "skills",
                    []
                ),

            "candidate_education":
                candidate.get(
                    "education",
                    []
                ),

            "candidate_experience":
                candidate.get(
                    "experience",
                    []
                ),

            "match_score":
                score,

            "recommendation":
                recommendation,

            "matched_skills":
                result.get(
                    "matched_skills",
                    []
                ),

            "missing_skills":
                result.get(
                    "missing_skills",
                    []
                ),

            "strengths":
                result.get(
                    "strengths",
                    []
                ),

            "justification":
                result.get(
                    "justification",
                    ""
                )
        })


    ranked_candidates.sort(
        key=lambda candidate:
            candidate["match_score"],
        reverse=True
    )


    for index, candidate in enumerate(
        ranked_candidates,
        start=1
    ):

        candidate["rank"] = index


    shortlisted_count = sum(

        1

        for candidate
        in ranked_candidates

        if candidate[
            "recommendation"
        ] == "SHORTLIST"
    )


    return {

        "message":
            "Screening calculated successfully",

        "job_id":
            job_id,

        "job_title":
            job.get(
                "title",
                ""
            ),

        "total_candidates":
            len(
                ranked_candidates
            ),

        "shortlisted":
            shortlisted_count,

        "candidates":
            ranked_candidates
    }


@router.get("/my-results")
def get_my_screening_results(

    current_user=Depends(
        get_current_user
    )
):

    current_user_id = str(
        current_user["_id"]
    )


    candidate_query = {

        "$or": [

            {
                "uploaded_by":
                    current_user_id
            },

            {
                "owner_id":
                    current_user_id
            },

            {
                "uploaded_by":
                    current_user["_id"]
            },

            {
                "owner_id":
                    current_user["_id"]
            }

        ]
    }


    candidates = list(
        candidates_collection.find(
            candidate_query
        )
    )


    candidate_ids = [

        candidate["_id"]

        for candidate in candidates
    ]


    if not candidate_ids:

        return {

            "total":
                0,

            "results":
                []
        }


    results = list(

        screening_results_collection
        .find(
            {
                "candidate_id":
                    {
                        "$in":
                            candidate_ids
                    }
            }
        )
        .sort(
            "match_score",
            -1
        )
    )


    formatted_results = []


    for result in results:

        candidate = (
            candidates_collection
            .find_one(
                {
                    "_id":
                        result[
                            "candidate_id"
                        ]
                }
            )
        )


        if not candidate:

            continue


        if not candidate_belongs_to_user(
            candidate,
            current_user
        ):

            continue


        job = (
            jobs_collection
            .find_one(
                {
                    "_id":
                        result[
                            "job_id"
                        ]
                }
            )
        )


        formatted_results.append(
            format_candidate_result(
                result,
                candidate,
                job
            )
        )


    return {

        "total":
            len(
                formatted_results
            ),

        "results":
            formatted_results
    }


@router.get("/results")
def get_screening_results(

    current_user=Depends(
        require_roles(
            "ADMIN",
            "RECRUITER"
        )
    )
):

    results = list(

        screening_results_collection
        .find({})
        .sort(
            "match_score",
            -1
        )
    )


    formatted_results = []


    for result in results:

        candidate = (
            candidates_collection
            .find_one(
                {
                    "_id":
                        result[
                            "candidate_id"
                        ]
                }
            )
        )


        job = (
            jobs_collection
            .find_one(
                {
                    "_id":
                        result[
                            "job_id"
                        ]
                }
            )
        )


        if not candidate:

            continue


        formatted_result = (
            format_candidate_result(
                result,
                candidate,
                job
            )
        )


        formatted_results.append(
            formatted_result
        )


    return {

        "total":
            len(
                formatted_results
            ),

        "results":
            formatted_results
    }


@router.get("/results/{result_id}")
def get_screening_result(

    result_id: str,

    current_user=Depends(
        get_current_user
    )
):

    try:

        result = (
            screening_results_collection
            .find_one(
                {
                    "_id":
                        ObjectId(
                            result_id
                        )
                }
            )
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid screening result ID"
        )


    if not result:

        raise HTTPException(
            status_code=404,
            detail="Screening result not found"
        )


    candidate = (
        candidates_collection
        .find_one(
            {
                "_id":
                    result[
                        "candidate_id"
                    ]
            }
        )
    )


    if not candidate:

        raise HTTPException(
            status_code=404,
            detail="Candidate not found"
        )


    user_role = str(
        current_user.get(
            "role",
            "viewer"
        )
    ).lower()


    if user_role == "viewer":

        if not candidate_belongs_to_user(
            candidate,
            current_user
        ):

            raise HTTPException(
                status_code=403,
                detail="You can only view your own screening results"
            )


    elif user_role not in [
        "admin",
        "recruiter"
    ]:

        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )


    job = (
        jobs_collection
        .find_one(
            {
                "_id":
                    result[
                        "job_id"
                    ]
            }
        )
    )


    return format_candidate_result(
        result,
        candidate,
        job
    )


@router.get("/job/{job_id}")
def get_job_screening_results(

    job_id: str,

    current_user=Depends(
        get_current_user
    )
):

    user_role = str(
        current_user.get(
            "role",
            "viewer"
        )
    ).lower()


    try:

        job_object_id = ObjectId(
            job_id
        )

    except Exception:

        raise HTTPException(
            status_code=400,
            detail="Invalid job ID"
        )


    job = (
        jobs_collection
        .find_one(
            {
                "_id":
                    job_object_id
            }
        )
    )


    if not job:

        raise HTTPException(
            status_code=404,
            detail="Job description not found"
        )


    if user_role == "viewer":

        current_user_id = str(
            current_user["_id"]
        )


        candidate_query = {

            "$or": [

                {
                    "uploaded_by":
                        current_user_id
                },

                {
                    "owner_id":
                        current_user_id
                },

                {
                    "uploaded_by":
                        current_user["_id"]
                },

                {
                    "owner_id":
                        current_user["_id"]
                }

            ]
        }


        candidates = list(
            candidates_collection.find(
                candidate_query
            )
        )


        candidate_ids = [

            candidate["_id"]

            for candidate in candidates
        ]


        if not candidate_ids:

            return {

                "job_id":
                    job_id,

                "job_title":
                    job.get(
                        "title",
                        ""
                    ),

                "total":
                    0,

                "results":
                    []
            }


        results = list(

            screening_results_collection
            .find(
                {
                    "job_id":
                        job_object_id,

                    "candidate_id":
                        {
                            "$in":
                                candidate_ids
                        }
                }
            )
            .sort(
                "match_score",
                -1
            )
        )


        formatted_results = []


        for result in results:

            candidate = (
                candidates_collection
                .find_one(
                    {
                        "_id":
                            result[
                                "candidate_id"
                            ]
                    }
                )
            )


            if not candidate:

                continue


            if not candidate_belongs_to_user(
                candidate,
                current_user
            ):

                continue


            formatted_results.append(
                format_candidate_result(
                    result,
                    candidate,
                    job
                )
            )


        return {

            "job_id":
                job_id,

            "job_title":
                job.get(
                    "title",
                    ""
                ),

            "total":
                len(
                    formatted_results
                ),

            "results":
                formatted_results
        }


    if user_role not in [
        "admin",
        "recruiter"
    ]:

        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )


    results = list(

        screening_results_collection
        .find(
            {
                "job_id":
                    job_object_id
            }
        )
        .sort(
            "match_score",
            -1
        )
    )


    formatted_results = []


    for index, result in enumerate(
        results,
        start=1
    ):

        candidate = (
            candidates_collection
            .find_one(
                {
                    "_id":
                        result[
                            "candidate_id"
                        ]
                }
            )
        )


        if not candidate:

            continue


        formatted_result = (
            format_candidate_result(
                result,
                candidate,
                job,
                index
            )
        )


        formatted_results.append(
            formatted_result
        )


    for index, candidate in enumerate(
        formatted_results,
        start=1
    ):

        candidate["rank"] = index


    return {

        "job_id":
            job_id,

        "job_title":
            job.get(
                "title",
                ""
            ),

        "total":
            len(
                formatted_results
            ),

        "results":
            formatted_results
    }

@router.get("/job/{job_id}/export")
def export_final_shortlist(
    job_id: str,
    current_user=Depends(
        require_roles(
            "ADMIN",
            "RECRUITER"
        )
    )
):
    try:
        job_object_id = ObjectId(
            job_id
        )
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Invalid job ID"
        )

    job = jobs_collection.find_one(
        {
            "_id": job_object_id
        }
    )

    if not job:
        raise HTTPException(
            status_code=404,
            detail="Job description not found"
        )

    results = list(
        screening_results_collection
        .find(
            {
                "job_id": job_object_id
            }
        )
        .sort(
            "match_score",
            -1
        )
    )

    if not results:
        raise HTTPException(
            status_code=404,
            detail="No screening results found for this job"
        )

    shortlist_results = [
        result
        for result in results
        if str(
            result.get(
                "recommendation",
                ""
            )
        ).upper() == "SHORTLIST"
    ]

    review_results = [
        result
        for result in results
        if str(
            result.get(
                "recommendation",
                ""
            )
        ).upper() == "REVIEW"
    ]

    reject_results = [
        result
        for result in results
        if str(
            result.get(
                "recommendation",
                ""
            )
        ).upper() == "REJECT"
    ]

    workbook = Workbook()

    shortlist_sheet = workbook.active
    shortlist_sheet.title = "Shortlist"

    review_sheet = workbook.create_sheet(
        "Review"
    )

    reject_sheet = workbook.create_sheet(
        "Reject"
    )

    headers = [
        "Rank",
        "Candidate",
        "Email",
        "Phone",
        "Match Score",
        "Matched Skills",
        "Missing Skills",
        "Strengths",
        "Justification",
        "Recommendation"
    ]

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1E3A8A"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_side = Side(
        style="thin",
        color="D1D5DB"
    )

    border = Border(
        bottom=thin_side
    )

    column_widths = {
        "A": 10,
        "B": 30,
        "C": 35,
        "D": 20,
        "E": 15,
        "F": 40,
        "G": 40,
        "H": 45,
        "I": 65,
        "J": 20
    }


    def setup_sheet(
        worksheet
    ):
        worksheet.append(
            headers
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = border

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        worksheet.freeze_panes = "A2"

        worksheet.row_dimensions[
            1
        ].height = 30


    def add_results(
        worksheet,
        category_results
    ):
        for index, result in enumerate(
            category_results,
            start=1
        ):
            candidate = (
                candidates_collection
                .find_one(
                    {
                        "_id":
                            result[
                                "candidate_id"
                            ]
                    }
                )
            )

            if not candidate:
                continue

            matched_skills = result.get(
                "matched_skills",
                []
            )

            missing_skills = result.get(
                "missing_skills",
                []
            )

            strengths = result.get(
                "strengths",
                []
            )

            worksheet.append(
                [
                    index,

                    candidate.get(
                        "name",
                        "Unknown Candidate"
                    ),

                    candidate.get(
                        "email",
                        ""
                    ),

                    candidate.get(
                        "phone",
                        ""
                    ),

                    result.get(
                        "match_score",
                        0
                    ),

                    ", ".join(
                        matched_skills
                    )
                    if isinstance(
                        matched_skills,
                        list
                    )
                    else str(
                        matched_skills
                    ),

                    ", ".join(
                        missing_skills
                    )
                    if isinstance(
                        missing_skills,
                        list
                    )
                    else str(
                        missing_skills
                    ),

                    "\n".join(
                        strengths
                    )
                    if isinstance(
                        strengths,
                        list
                    )
                    else str(
                        strengths
                    ),

                    result.get(
                        "justification",
                        ""
                    ),

                    result.get(
                        "recommendation",
                        "REVIEW"
                    )
                ]
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


    setup_sheet(
        shortlist_sheet
    )

    setup_sheet(
        review_sheet
    )

    setup_sheet(
        reject_sheet
    )


    add_results(
        shortlist_sheet,
        shortlist_results
    )

    add_results(
        review_sheet,
        review_results
    )

    add_results(
        reject_sheet,
        reject_results
    )


    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)


    safe_job_title = (
        job.get(
            "title",
            "job"
        )
        .replace(
            " ",
            "_"
        )
        .replace(
            "/",
            "_"
        )
        .replace(
            "\\",
            "_"
        )
    )


    filename = (
        f"{safe_job_title}_"
        f"screening_results.xlsx"
    )


    return StreamingResponse(
        output,
        media_type=
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={
            "Content-Disposition":
                f'attachment; filename="{filename}"'
        }
    )